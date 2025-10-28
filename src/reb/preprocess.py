
# ------------------------------------------------------------------------ #
#                                  MODULES                                 #
# ------------------------------------------------------------------------ #

# built-in modules
from pathlib  import Path
from openpyxl import load_workbook
from typing   import (
    Any     ,
    List    ,
    Optional,
)
from datetime import datetime

# third-party modules

# local modules
from epsimple.core.construction import (
    SurfaceConstruction     ,
    SurfaceType             ,
    SurfaceBoundaryCondition,
)
from epsimple.core.model import address_to_weather

# ---------------------------------------------------------------------------- #
#                           PHIKO EXCEL PREPROCESSING                          #
# ---------------------------------------------------------------------------- #

#%%
def load_excel(file_path: str, data_only: bool = False): # data_only를 받을 수 있도록 수정됨
    """엑셀 로드 (서식 유지). 반환: openpyxl Workbook"""
    wb = load_workbook(file_path, data_only=data_only)
    return wb

def _headers_of(ws, header_row: int):
    return [cell.value for cell in ws[header_row]]

def _col_index_by_name(ws, header_row: int, col_name: str) -> int:
    headers = _headers_of(ws, header_row)
    if col_name not in headers:
        raise ValueError(f"컬럼 '{col_name}' 없음. 시트 헤더: {headers}")
    return headers.index(col_name) + 1  # 1-based index

def drop_rows_inplace(
    wb,
    sheet_name: str,
    col_name: str,
    values: List[Any],
    header_row: int = 1,
    normalize: bool = True,
    *,
    verbose:bool=True,
    ):
    """
    주어진 워크북(wb)에서, sheet_name 시트의 col_name 컬럼이
    values 목록과 일치하는 행들을 '서식 유지'한 채로 삭제 (메모리에서만 변경).
    """
    def norm(v):
        if not normalize:
            return v
        if v is None:
            return None
        return str(v).strip()

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}' 없음. 현재 시트들: {wb.sheetnames}")

    ws = wb[sheet_name]
    col_idx = _col_index_by_name(ws, header_row, col_name)
    targets = set(norm(v) for v in values)

    rows_to_delete = []
    for row in ws.iter_rows(min_row=header_row + 1):
        cell_val = norm(row[col_idx - 1].value)
        if cell_val in targets:
            rows_to_delete.append(row[0].row)

    rows_to_delete.sort(reverse=True)
    for r in rows_to_delete:
        ws.delete_rows(r, 1)

    if verbose:
        print(f"[삭제] 시트='{sheet_name}', 컬럼='{col_name}', 삭제행수={len(rows_to_delete)}")

def multiply_column_inplace(
    wb,
    sheet_name: str,
    col_name: str,
    factor: float = 1000.0,
    header_row: int = 1,
    coerce_numeric: bool = True,  
    skip_formula: bool = True,
    *,
    verbose:bool=True,
) -> int:
    """
    시트 `sheet_name`의 컬럼 `col_name` 전체 값을 factor 배로 곱하여 덮어씀.
    - 숫자 타입(int/float)은 바로 곱함
    - 문자열 숫자는 coerce_numeric=True이면 숫자로 변환 후 곱함 (쉼표 제거)
    - 수식은 기본적으로 건드리지 않음 (skip_formula=True)

    Returns
    -------
    int : 변경된 셀 개수
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}' 없음. 현재 시트들: {wb.sheetnames}")

    ws = wb[sheet_name]
    col_idx = _col_index_by_name(ws, header_row, col_name)
    changed = 0

    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[col_idx - 1]
        val = cell.value

        if val is None:
            continue

        # 수식은 패스
        if skip_formula and isinstance(val, str) and val.startswith("="):
            continue

        if isinstance(val, (int, float)):
            cell.value = val * factor
            changed += 1
            continue

        if coerce_numeric and isinstance(val, str):
            s = val.replace(",", "").strip()
            try:
                num = float(s)
            except ValueError:
                continue
            # 원래 정수 형태였다면 int로 유지
            cell.value = int(num * factor) if s.isdigit() and float(s).is_integer() else (num * factor)
            changed += 1

    if verbose:
        print(f"[곱셈] 시트='{sheet_name}', 컬럼='{col_name}', factor={factor}, 변경셀수={changed}")
    return changed

def convert_formulas_in_column_to_values(
    wb_main,
    wb_values,
    sheet_name: str,
    col_name: str,
    header_row: int = 1,
    *,
    verbose:bool=True,
):
    """
    특정 '컬럼(col_name)'에 포함된 모든 수식을 계산된 값으로 덮어씁니다.
    wb_main: 실제 수정이 일어나는 메인 워크북 (수식 보존)
    wb_values: 값 조회용으로만 사용하는 워크북 (수식이 값으로 변환됨)
    """
    if sheet_name not in wb_main.sheetnames:
        print(f"[경고] 시트 '{sheet_name}'가 없어 수식 변환을 건너뜁니다.")
        return

    ws_main = wb_main[sheet_name]
    ws_values = wb_values[sheet_name]
    
    try:
        col_idx = _col_index_by_name(ws_main, header_row, col_name)
    except ValueError as e:
        print(f"[경고] {e}. '{sheet_name}' 시트의 수식 변환 작업을 건너뜁니다.")
        return

    converted_count = 0
    # iter_rows를 사용하여 메인 워크북과 값 조회용 워크북의 행을 동시에 순회
    for row_main, row_values in zip(
        ws_main.iter_rows(min_row=header_row + 1),
        ws_values.iter_rows(min_row=header_row + 1)
    ):
        cell_main = row_main[col_idx - 1]

        # 메인 워크북의 셀이 수식('f') 타입일 때만 값으로 덮어씀
        if cell_main.data_type == 'f':
            # 값 조회용 워크북에서 동일한 위치의 셀 값을 가져옴
            cell_with_value = row_values[col_idx - 1]
            cell_main.value = cell_with_value.value
            converted_count += 1
    
    if verbose and converted_count > 0:
        print(f"[변환] 시트='{sheet_name}', 컬럼='{col_name}'의 수식 {converted_count}개를 값으로 변경했습니다.")
        
def convert_이전레이어(wb, verbose:bool=True) -> None:
    
    sheet = wb["구조체_면"]
    
    address = wb["건물정보"].cell(2,3).value
    vintage = datetime(*[int(v) for v in wb["건물정보"].cell(2,6).value.split("-")])
    _, climate, _, _ = address_to_weather(address, vintage)
    
    for 구조체_row in sheet.iter_rows(min_row=2):
        
        if all(cell.value is None for cell in 구조체_row):
            break
        
        if 구조체_row[1].value == '&SPECIAL&이전레이어&':
            
            구조체이름 = 구조체_row[0].value
            원래값들 = [cell.value for cell in 구조체_row[3:11] if cell.value is not None]
            
            for 면_row in wb["면"].iter_rows(min_row=2):
                
                if 면_row[7].value == 구조체이름:
                    면이름 = 면_row[0].value
                    추가할면별구조체이름 =f"{구조체이름}for{면이름}" 
                    면_row[7].value = 추가할면별구조체이름
                    
                    if 면_row[0].value is None:
                        break
                    
                    현재구조체개수 = len([sheet.cell(count+1,1) for count in range(1, sheet.max_row) if sheet.cell(count+1,1).value is not None])
                    surface_type = SurfaceType(면_row[2].value)
                    surface_boundary = SurfaceBoundaryCondition(면_row[3].value)
                    
                    이전surface = SurfaceConstruction.get_regulated_construction(
                        vintage,
                        surface_type,
                        surface_boundary,
                        climate,
                    )
                    
                    현재재료 = [wb["재료"].cell(count+1,1).value for count in range(1, wb["재료"].max_row) if wb["재료"].cell(count+1,1).value is not None]
                    현재재료개수 = len(현재재료)
                    for layer in 이전surface.layers:
                        추가할재료이름 = layer[0].ID

                        if 추가할재료이름 not in 현재재료:
                            wb["재료"].cell(현재재료개수+2,1).value = 추가할재료이름
                            wb["재료"].cell(현재재료개수+2,2).value = layer[0].conductivity
                            wb["재료"].cell(현재재료개수+2,3).value = layer[0].density
                            wb["재료"].cell(현재재료개수+2,4).value = layer[0].specific_heat
                    
                    sheet.cell(현재구조체개수+2, 1).value = 추가할면별구조체이름
                    for idx, layer in enumerate(이전surface.layers):
                        sheet.cell(현재구조체개수+2, 2*(idx+1)  ).value = layer[0].ID
                        sheet.cell(현재구조체개수+2, 2*(idx+1)+1).value = layer[1] * 1000
                    
                    for idx2, value in enumerate(원래값들):
                        sheet.cell(현재구조체개수+2, 2*(idx+1)+2+idx2).value = value
                    
    drop_rows_inplace(wb, sheet_name="구조체_면", col_name="레이어1_재료", values=['&SPECIAL&이전레이어&'], verbose=verbose)        


def replace_typo(wb) -> None:
    
    # 오타 1
    for idx in range(wb["실"].max_row):
        if wb["실"].cell(idx+1, 4).value == "회의실 및 세미나실":
            wb["실"].cell(idx+1, 4).value = "회의 및 세미나실"
        if wb["실"].cell(idx+1, 4).value == "강의실(대학)":
            wb["실"].cell(idx+1, 4).value = "강의실"
        if wb["실"].cell(idx+1, 4).value == "매장(상점/백화점)":
            wb["실"].cell(idx+1, 4).value = "매장"
        if wb["실"].cell(idx+1, 4).value == "전시실(전시관/박물관)":
            wb["실"].cell(idx+1, 4).value = "전시실"
        if wb["실"].cell(idx+1, 4).value == "열람실(도서관)":
            wb["실"].cell(idx+1, 4).value = "열람실"
    
    return 
    
def save_excel(
    wb,
    output_filepath: Optional[str] = None,
    *,
    original_filepath: Optional[str] = None,
    suffix: str = "_preprocess",
    verbose:bool = True
):
    """
    워크북 저장만 담당.
    - output_path가 주어지면 그 경로에 저장
    - 없으면 original_path의 파일명에 '{suffix}'를 붙여 저장  <- 여기만 변경
      (예: foo.xlsx -> foo_preprocess.xlsx)
    """
    if output_filepath is None and original_filepath is None:
        raise ValueError("output_path 또는 original_path 중 하나는 지정해야 합니다.")

    if output_filepath is None:
        p = Path(original_filepath)                                  # <- 여기만 변경
        output_filepath = str(p.with_name(f"{p.stem}{suffix}{p.suffix}"))  # <- 여기만 변경

    wb.save(output_filepath)
    if verbose:
        print(f"[저장] {output_filepath}")
    
    return output_filepath


def process_excel_file(
    file_path:str,
    *,
    suffix         :str="_preprocess",
    output_filepath:str=None        ,
    verbose:bool=True,
    ) -> str:
    
        """하나의 엑셀 파일에 대한 전체 전처리 작업을 수행합니다."""
    # try:
        # 1. 엑셀 파일 로드
        wb = load_excel(file_path)
        wb_values = load_excel(file_path, data_only=True)
        # 2. 행 삭제 작업
        drop_rows_inplace(wb, sheet_name="구조체_면", col_name="이름", values=["open"], verbose=verbose)
        drop_rows_inplace(wb, sheet_name="재료", col_name="이름", values=["공기층", r"&SPECIAL&이전레이어&"], verbose=verbose)
        
        # 3. 수식을 숫자로 변경
        convert_formulas_in_column_to_values( 
            wb_main=wb,
            wb_values=wb_values,
            sheet_name="건물정보",
            col_name="north_axis [°]",
            verbose = verbose,
        ) # 예시: 이 컬럼에 있는 모든 수식을 값으로 변경
        convert_formulas_in_column_to_values(
            wb_main=wb,
            wb_values=wb_values,
            sheet_name="면",
            col_name="면적 [m2]",
            verbose = verbose,
        ) # 예시: 이 컬럼에 있는 모든 수식을 값으로 변경
        convert_formulas_in_column_to_values(
            wb_main=wb,
            wb_values=wb_values,
            sheet_name="개구부",
            col_name="면적 [m2]",
            verbose = verbose, 
        ) # 예시: 이 컬럼에 있는 모든 수식을 값으로 변경
        convert_formulas_in_column_to_values(
            wb_main=wb,
            wb_values=wb_values,
            sheet_name="생산설비",
            col_name="냉방COP [W/W]",
            verbose = verbose, 
        ) # 예시: 이 컬럼에 있는 모든 수식을 값으로 변경
        convert_formulas_in_column_to_values(
            wb_main=wb,
            wb_values=wb_values,
            sheet_name="생산설비",
            col_name="난방COP [W/W]",
            verbose = verbose, 
        ) # 예시: 이 컬럼에 있는 모든 수식을 값으로 변경
        
        # 4. 특정 열 값 일괄 곱셈
        multiply_column_inplace(
            wb,
            sheet_name="재료",
            col_name="비열 [J/kg·K]",
            factor=1000,
            verbose=False)

        # 오타 수정
        replace_typo(wb)
        
        # 이전레이어 처리
        convert_이전레이어(
            wb,
            verbose=verbose
        )
        
        # 5. 결과 저장
        if output_filepath is None:
            output_filepath = save_excel(wb, original_filepath=file_path, suffix=suffix, verbose=verbose)
        else:
            output_filepath = save_excel(wb, output_filepath, verbose=verbose)
        
        return output_filepath


